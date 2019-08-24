from infoScraper import *
from collections import Counter
import os
import openpyxl
from openpyxl.utils import get_column_letter
import sys
sys.path.append("..")

def generateBOM(srcFile=defaultFile(),destFile="output/UnnamedBOM"):
    #- Gets Part info and calculates Quantity of Hardware types as well as total price for each type
    if not os.path.exists('output'):
        os.makedirs('output')
    infoRows = scrapePartData(srcFile)
    itemCount = Counter([tuple(row) for row in infoRows])
    infoRows = [[row[0],float(row[1]),int(itemCount[row]),float(row[1]*itemCount[row]),row[2]] for row in itemCount.keys()]
    OverallCost = sum([float(row[3]) for row in infoRows])
    #- Generates an excel workbook and fills its active worksheet with Hardware type info
    wb = openpyxl.Workbook()
    ws = wb.active
    bomHeaders = ["Name","Price","Quantity","Total","URL"]
    for col in range(1,len(bomHeaders)+1):
        ws[get_column_letter(col)+"1"].value = bomHeaders[col-1]
    for row in range(len(infoRows)):
        for column in range(len(bomHeaders)):
            ws[get_column_letter(column+1)+str(row+2)].value = infoRows[row][column]
    ws["A{}".format(len(infoRows)+2)].value = "OverallCost"
    ws["{}{}".format(get_column_letter(4),len(infoRows)+2)].value = OverallCost
    wb.save("{}.xlsx".format(destFile))

generateBOM()